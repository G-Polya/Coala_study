import requests
from bs4 import BeautifulSoup
import openpyxl


keyword = input('검색어를 입력해주세요: ')

try:
    wb = openpyxl.load_workbook('naver_News.xlsx')
    sheet = wb.active
    print('불러오기 완료')
except:
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(['제목', '언론사'])
    print('새로운 파일을 만들었습니다')
# wb = openpyxl.Workbook()

for page in range(1,101,10):
    raw = requests.get('https://search.naver.com/search.naver?where=news&query='+keyword+'&start='+str(page),
                       headers = {'User-Agent':'Mozilla/5.0'})

    html = BeautifulSoup(raw.text,'html.parser')

    # 컨테이너 : ul.type01 > li
    # 기사제목 : a._sp_each_title
    # 언론사 : span._sp_each_source

    #1. 컨테이너 수집
    articles = html.select('ul.type01 > li')

    #2. 기사데이터 수집
    for ar in articles:
        title = ar.select_one('a._sp_each_title').text
        source = ar.select_one('span._sp_each_source').text

        print(title, source)
        print('-' * 50)

        sheet.append([title, source])

wb.save('naver_News.xlsx')